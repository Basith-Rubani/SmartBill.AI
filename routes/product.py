from flask import Blueprint, render_template, request, jsonify
from flask_login import login_required
from models import db, Product
import io
import openpyxl
import re
    
from flask import send_file
import tempfile

bp = Blueprint("product", __name__, url_prefix="/products")  # ✅ fixed


# ---------- Helpers ----------
def _to_dict(p: Product):
    return {
        "id": p.id,
        "name": p.name,
        "category": getattr(p, "category", "") or "",
        "price": float(p.price),
        "stock": int(p.stock),
        "gst": (
            float(getattr(p, "gst", 0.18))
            if getattr(p, "gst", None) is not None
            else None
        ),
    }


# ---------- Pages ----------
@bp.route("/", methods=["GET"])
@login_required
def page_products():
    return render_template("product.html")


# ---------- APIs ----------
@bp.route("/api", methods=["GET"])
@login_required
def api_list():
    q = request.args.get("q", "").strip()
    sort = request.args.get("sort", "name")

    query = Product.query

    if q:
        q_like = f"%{q}%"

        if q.isdigit():
            query = query.filter(
                (Product.id == int(q))
                | Product.name.ilike(q_like)
                | getattr(Product, "category", Product.name).ilike(q_like)
            )
        else:
            query = query.filter(
                Product.name.ilike(q_like)
                | getattr(Product, "category", Product.name).ilike(q_like)
            )

    desc = False
    if sort.startswith("-"):
        desc = True
        sort = sort[1:]

    order_col = Product.name
    if sort == "price":
        order_col = Product.price
    elif sort == "stock":
        order_col = Product.stock

    if desc:
        order_col = order_col.desc()

    products = query.order_by(order_col).all()

    return jsonify({
        "products": [_to_dict(p) for p in products]
    })


@bp.route("/api", methods=["POST"])
@login_required
def api_create():
    data = request.get_json() or {}

    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Product name required"}), 400

    try:
        price = float(data.get("price", 0))
        stock = int(data.get("stock", 0))
    except Exception:
        return jsonify({"error": "Invalid numeric values"}), 400

    category = data.get("category", "") or ""
    gst = data.get("gst", None)

    p = Product(
        name=name,
        price=price,
        stock=stock,
        category=category,
    )

    if hasattr(p, "gst") and gst is not None:
        try:
            p.gst = float(gst)
        except Exception:
            pass

    db.session.add(p)
    db.session.commit()

    return jsonify({
        "message": "Created",
        "product": _to_dict(p),
    }), 201


@bp.route("/api/<int:product_id>", methods=["PUT"])  # ✅ fixed syntax
@login_required
def api_update(product_id):
    p = Product.query.get_or_404(product_id)
    data = request.get_json() or {}

    if "name" in data:
        p.name = str(data["name"]).strip()

    if "category" in data:
        p.category = data["category"] or ""

    if "price" in data:
        try:
            p.price = float(data["price"])
        except Exception:
            return jsonify({"error": "Invalid price"}), 400

    if "stock" in data:
        try:
            p.stock = int(data["stock"])
        except Exception:
            return jsonify({"error": "Invalid stock"}), 400

    if "gst" in data and hasattr(p, "gst"):
        try:
            p.gst = float(data["gst"])
        except Exception:
            pass

    db.session.commit()

    return jsonify({
        "message": "Updated",
        "product": _to_dict(p),
    })


@bp.route("/api/<int:product_id>", methods=["DELETE"])  # ✅ fixed syntax
@login_required
def api_delete(product_id):
    p = Product.query.get_or_404(product_id)

    db.session.delete(p)
    db.session.commit()

    return jsonify({"message": "Deleted"})

@bp.route("/api/import", methods=["POST"])
@login_required
def api_import_excel():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]

    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files supported"}), 400

    try:
        in_memory = io.BytesIO(file.read())
        workbook = openpyxl.load_workbook(in_memory)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "Excel file is empty"}), 400

        headers = [str(h).strip().lower() for h in rows[0]]

        # Auto detect columns
        def find_col(possible_names):
            for name in possible_names:
                if name in headers:
                    return headers.index(name)
            return None

        col_name = find_col(["name", "product", "product name", "item"])
        col_category = find_col(["category"])
        col_price = find_col(["price", "rate", "amount"])
        col_stock = find_col(["stock", "qty", "quantity"])
        col_gst = find_col(["gst", "tax"])

        if col_name is None:
            return jsonify({"error": "Product name column not found"}), 400

        added = 0
        updated = 0

        for row in rows[1:]:
            if not row[col_name]:
                continue

            name = str(row[col_name]).strip()

            # Safe parsing function
            def parse_float(val):
                if val is None:
                    return 0.0
                if isinstance(val, (int, float)):
                    return float(val)
                val = re.sub(r"[^\d.]", "", str(val))
                return float(val) if val else 0.0

            def parse_int(val):
                if val is None:
                    return 0
                if isinstance(val, int):
                    return val
                val = re.sub(r"[^\d]", "", str(val))
                return int(val) if val else 0

            category = (
                str(row[col_category]).strip()
                if col_category is not None and row[col_category]
                else ""
            )

            price = (
                parse_float(row[col_price])
                if col_price is not None
                else 0
            )

            stock = (
                parse_int(row[col_stock])
                if col_stock is not None
                else 0
            )

            gst = (
                parse_float(row[col_gst])
                if col_gst is not None
                else 0
            )

            # DUPLICATE CHECK (by name)
            existing = Product.query.filter(
                Product.name.ilike(name)
            ).first()

            if existing:
                existing.category = category
                existing.price = price
                existing.stock = stock
                if hasattr(existing, "gst"):
                    existing.gst = gst
                updated += 1
            else:
                new_product = Product(
                    name=name,
                    category=category,
                    price=price,
                    stock=stock,
                )
                if hasattr(new_product, "gst"):
                    new_product.gst = gst
                db.session.add(new_product)
                added += 1

        db.session.commit()

        return jsonify({
            "message": f"Import complete",
            "added": added,
            "updated": updated
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/template", methods=["GET"])
@login_required
def download_template():
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.append(["name", "category", "price", "stock", "gst"])
    sheet.append(["Sample Product", "General", 100, 50, 18])

    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    workbook.save(temp.name)

    return send_file(
        temp.name,
        as_attachment=True,
        download_name="product_import_template.xlsx"
    )
